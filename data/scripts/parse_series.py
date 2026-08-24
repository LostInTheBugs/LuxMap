#!/usr/bin/env python3
"""Build multi-year series per commune → data/processed/series.json.

Sources:
  - vente-appartement-2010-2025.xlsx / vente-maison-2010-2025.xlsx
    (one sheet per year, 2010-2025; col 2 = commune, col 5 = €/m²)
  - loyers-annonces-apparts-2009-2025.xls (same layout, 2009-2025)
  - chomage.json (STATEC LUSTAT DF_X026): per-year C6 rates per commune

Output: {"prix_appart": {year: {lau2: value}}, ...} joined on LAU2 via the
commune name (norm()) like build_data.py.

Run with: uv run --with xlrd --with openpyxl python3 data/scripts/parse_series.py
"""
import json
import os

import xlrd

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.normpath(os.path.join(HERE, "..", "raw"))
OUT = os.path.normpath(os.path.join(HERE, "..", "processed"))

BLACKLIST = {"luxembourg ville", "luxembourg"}


def norm(s: str) -> str:
    return s.strip().lower().replace("-", " ").replace("'", " ").replace("/", " ").replace("  ", " ")


def load_communes() -> dict[str, str]:
    """norm(name) -> lau2 for the 100 current communes."""
    with open(os.path.join(RAW, "limadmin.geojson"), encoding="utf-8") as f:
        gj = json.load(f)
    return {
        norm(f["properties"]["COMMUNE"]): f["properties"]["LAU2"]
        for f in gj["communes"]["features"]
    }


def parse_rows(sheet) -> dict[str, float]:
    out: dict[str, float] = {}
    for r in range(sheet.nrows):
        commune = str(sheet.cell_value(r, 2)).strip()
        if not commune or commune == "Commune":
            continue
        val = sheet.cell_value(r, 5)
        if isinstance(val, str) and val.strip() in ("*", ""):
            continue
        try:
            out[commune] = round(float(val), 2)
        except (TypeError, ValueError):
            continue
    return out


def parse_xls_series(path: str) -> dict[str, dict[str, float]]:
    wb = xlrd.open_workbook(path)
    out: dict[str, dict[str, float]] = {}
    for name in wb.sheet_names():
        try:
            year = str(int(float(name)))
        except (TypeError, ValueError):
            continue
        if len(year) != 4:
            continue
        out[year] = parse_rows(wb.sheet_by_name(name))
    return out


def parse_xlsx_series(path: str) -> dict[str, dict[str, float]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[str, dict[str, float]] = {}
    for name in wb.sheetnames:
        if not (len(name) == 4 and name.isdigit()):
            continue
        sh = wb[name]
        rows: dict[str, float] = {}
        for r in range(1, sh.max_row + 1):
            commune = sh.cell(r, 3).value
            if not commune or str(commune).strip() == "Commune":
                continue
            val = sh.cell(r, 6).value
            if val is None or (isinstance(val, str) and val.strip() in ("*", "")):
                continue
            try:
                rows[str(commune).strip()] = round(float(val), 2)
            except (TypeError, ValueError):
                continue
        out[name] = rows
    return out


def parse_chomage_series() -> dict[str, dict[str, float]]:
    """{year: {commune_name: rate}} from the LUSTAT SDMX-JSON dump."""
    with open(os.path.join(RAW, "chomage.json"), encoding="utf-8") as f:
        d = json.load(f)
    st = d["structure"]
    var_dim = st["dimensions"]["series"][0]
    spec_dim = st["dimensions"]["series"][1]
    rate_idx = next(i for i, v in enumerate(var_dim["values"]) if v["id"] == "C6")
    periods = [v["id"] for v in st["dimensions"]["observation"][0]["values"]]
    spec_names = {v["id"]: v["name"] for v in spec_dim["values"]}
    out: dict[str, dict[str, float]] = {}
    for sk, sv in d["dataSets"][0]["series"].items():
        vi, si, _ = (int(x) for x in sk.split(":"))
        if vi != rate_idx:
            continue
        name = spec_names[spec_dim["values"][si]["id"]]
        for i, o in sv.get("observations", {}).items():
            v = o[0]
            if v is None or v == "" or int(i) >= len(periods):
                continue
            out.setdefault(periods[int(i)], {})[name] = round(float(v), 1)
    return out


def join_by_name(series: dict[str, dict[str, float]], lau2_by_norm: dict[str, str]) -> dict[str, dict[str, str]]:
    """{year: {lau2: value}}"""
    out: dict[str, dict[str, str]] = {}
    for year, by_name in series.items():
        joined = {}
        for name, v in by_name.items():
            n = norm(name)
            if n in BLACKLIST or n not in lau2_by_norm:
                continue
            joined[lau2_by_norm[n]] = v
        out[year] = joined
    return out


def transpose_lau2_series(data: dict[str, dict[str, float]]) -> dict[str, dict[str, float]]:
    """{lau2: {year: value}} → {year: {lau2: value}} (outputs of parse_lustat2.py)."""
    out: dict[str, dict[str, float]] = {}
    for lau2, by_year in data.items():
        for year, v in by_year.items():
            out.setdefault(year, {})[lau2] = v
    return out


def main() -> None:
    os.makedirs(OUT, exist_ok=True)
    lau2 = load_communes()

    with open(os.path.join(OUT, "population.json"), encoding="utf-8") as f:
        population = transpose_lau2_series(json.load(f))
    with open(os.path.join(OUT, "accidents.json"), encoding="utf-8") as f:
        accidents = transpose_lau2_series(json.load(f))
    with open(os.path.join(OUT, "natural.json"), encoding="utf-8") as f:
        natural = transpose_lau2_series(json.load(f))
    with open(os.path.join(OUT, "migration.json"), encoding="utf-8") as f:
        migration = transpose_lau2_series(json.load(f))

    series: dict[str, dict[str, dict[str, float]]] = {
        "prix_appart": join_by_name(parse_xlsx_series(os.path.join(RAW, "vente-appartement-2010-2025.xlsx")), lau2),
        "prix_maison": join_by_name(parse_xlsx_series(os.path.join(RAW, "vente-maison-2010-2025.xlsx")), lau2),
        "loyer_appart": join_by_name(parse_xls_series(os.path.join(RAW, "loyers-annonces-apparts-2009-2025.xls")), lau2),
        "chomage": join_by_name(parse_chomage_series(), lau2),
        "population": population,
        "accidents": accidents,
        "solde_naturel": natural,
        "solde_migratoire": migration,
    }

    with open(os.path.join(OUT, "series.json"), "w", encoding="utf-8") as f:
        json.dump(series, f, ensure_ascii=False)

    for key, by_year in series.items():
        years = sorted(by_year)
        cov = [f"{y}:{len(by_year[y])}" for y in years]
        print(f"{key}: {len(years)} ans — " + ", ".join(cov))


if __name__ == "__main__":
    main()
